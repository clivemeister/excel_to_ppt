## Function to extract the column data into a dict, indexed by row, ignoring empty cells
def get_column_to_dict(sheet,col,first_row,last_row):
    col_values=dict()
    for row in range(first_row_with_data,last_row_with_data):
        val=sheet.cell(row=row,column=col).value
        if (val is not None): col_values[row]=val.lower()
    return col_values

## Add up down all the columns of an MxN array, returning a 1xN array of the result
def sumColumns(m):
    return [sum(col) for col in zip(*m)][0]

from openpyxl import Workbook
import openpyxl
book = openpyxl.load_workbook('Insights.xlsx')
sheet = book.active

## Import the dates
from datetime import datetime
print("Reading the first column for dates")
row_date=dict()
first_row_with_data=0
last_row_with_data=0
#extract the dates into a dict, indexed by row, noting first and last rows with valid data
for row in range(1,2000):
    val=sheet.cell(row=row,column=1).value
    if (val is not None):
        #try and convert to date
        try:
            if (type(val) is datetime):
                dateval = val   # cell is already in date format, so just use it
            else:
                dateval = datetime.strptime(val,"%b %d, %Y")   # parse the values (in the format we seem to have) into a date
            row_date[row]=dateval
            if (first_row_with_data==0): first_row_with_data=row
            else: last_row_with_data=row
        except ValueError:
            print("Didn't find a parseable date in row ",row,": ",val)
        except TypeError:
            print("Type error on row ",row," for value ",val," of type ",type(val))

print("data runs from row ",first_row_with_data," to ",last_row_with_data)

## Extract non-empty values of column 13, "Want To Learn More About", into a dict, indexed by row
learn_more_about = get_column_to_dict(sheet,13,first_row_with_data,last_row_with_data)

## Extract non-empty values of col 18, "Action Items", into a dict, indexed by row
action_items = get_column_to_dict(sheet,18,first_row_with_data,last_row_with_data)

## Find all the text we will use from any row with given month's date
month_to_use = 8
rows_in_month=0
comments_to_search = []
for r,d in row_date.items():
    if (d.date().month == month_to_use):
        rows_in_month+=1
        newitem=""
        if (r in learn_more_about): newitem=learn_more_about[r]
        if (r in action_items): newitem=newitem+" "+action_items[r]
        if (newitem!=""): comments_to_search.append(newitem)
print("Found ",rows_in_month," rows, with ",len(comments_to_search)," cells to examine, for month ",month_to_use)

replace_text={
    "/": " ",
    "\n": " ",
    "big data": "big_data",
    "new stack": "new_stack",
    "flexible capacity": "flex_capacity",
    "gen 10": "gen10",
    "gen z": "gen-z",
    "cloud cruiser": "cloudcruiser",
    "store once": "storeonce",
    "smart city": "smart_city",
    "future city": "smart_city",
    "azure stack": "azurestack",
    "intelligent edge": "edgeline",
    "clear pass": "clearpass",
    "one view": "oneview",
    "the machine": "the_machine",
    "open stack": "openstack",
    "store once": "storeonce",
    "office 365": "office365",
    "hp financial services" "hpefs"
    "hpe financial services": "hpefs",
    "hpfs": "hpefs",
    "integrity": "superdome",
    "mobility": "wireless"
}

for i, val in enumerate(comments_to_search):
    updated_item=comments_to_search[i]
    #if (updated_item.find("the machine") != -1): print("found <the machine> in row ",i," - ",updated_item[:60])
    for k,v in replace_text.items():
        updated_item=updated_item.replace(k,v)
    comments_to_search[i]=updated_item

#print("First few texts found for month ",month_to_use,":\n",comments_to_search[:5])

#use this specific set of words as the vocab to look for
vocab=["iot","edgeline","smart_city", "big_data","sap","apollo", "saas", "analytics", "sgi", "networking", "wireless", "aruba",
       "arista", "clearpass", "naas", "skype", "meridian", "airwave","composable", "niara", "flex_capacity", "helion",
       "azurestack","synergy","moonshot","converged","hyperconverged","hybrid","docker","vdi","new_stack","cloudline",
       "cloudsystem","easyconnect","openstack","devops","3par","bura","simplivity","nimble","scality","storeonce",
       "oneview", "office365","pointnext","hpefs","the_machine","photonics","blockchain","pathfinder","gen10","gen-z",
       "cloudcruiser","blades","superdome"
      ]

#do the conversion to tf-idf, ignoring English stop words
from sklearn.feature_extraction.text import CountVectorizer

vectorizer = CountVectorizer(max_features=30,
                             min_df=2,
                             vocabulary=vocab,
                             binary=True   # This means the term freqency of any given word in an item is 0 (doesn't occur) or 1 (does, regardless of how often)
                            )
#This next line builds the documents x features matrix, using the vectorizer we've created
X = vectorizer.fit_transform(comments_to_search)

# Now turn the frequencies into a dict, mapping "word" to its frequency
dict_of_words_to_freq={}
col_sums = sumColumns(X)
for keyword,idx in vectorizer.vocabulary_.items():
    if col_sums[0,idx]>0:
        dict_of_words_to_freq[keyword]=col_sums[0,idx]
        #print(" keyword =",keyword," freq =",col_sums[0,idx])

import operator
sorted_list = sorted(dict_of_words_to_freq.items(), key=operator.itemgetter(1),reverse=True)
print("Keywords and counts found:", sorted_list)


##Build a wordcloud, using the wordcloud code from Andreas Mueller
# (to install, run "pip install wordcloud")
from wordcloud import WordCloud
import random

#Generate the wordcloud
wordcloud= WordCloud(width=2000,height=500,
                     prefer_horizontal=0.9,
                     relative_scaling=0.5,
                     max_font_size=144,
                     background_color="white"
                    ).generate_from_frequencies(dict_of_words_to_freq)
# Display the generated image
import matplotlib.pyplot as plt
plt.imshow(wordcloud,interpolation='bilinear')
plt.axis("off")
plt.figure()
plt.show()
